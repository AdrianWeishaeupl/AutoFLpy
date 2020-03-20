# -*- coding: utf-8 -*-
"""
Unit tests for the log_to_xlsx.py
code.

@author Adrian Weishaeupl (aw6g15@soton.ac.uk)
"""

from autoflpy.util import log_to_xlsx
import unittest
import os
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class TestLogToXlsx(unittest.TestCase):

    def setUp(self):
        # Create variables and assign directories before any test.
        base_path = os.path.join(os.path.dirname(__file__),
                                 "test_files") + os.sep
        # Tidies up the base path for python.
        self.base_path = base_path.replace(os.sep, "/")
        # Reads the test_input_file information
        with open((base_path + 'test_Input_File.json')) as file:
            self.data = json.load(file)
        # Sets dummy variables for testing to be used throughout.
        self.log_file_path = self.base_path + self.data["log_to_xlsx_input"][
                "log_file_name"]
        self.name_converter_file_path = self.base_path + \
            "test_name_converter_list.txt"
        self.data_sources_path = self.base_path + "test_data_sources.txt"
        self.excel_file_path = self.base_path
        self.excel_file_name = "test_xlsx"
        self.flight_date = self.data["log_to_xlsx_input"]["date"]
        self.flight_number = self.data["log_to_xlsx_input"]["flight_number"]
        self.weather_data = self.data["weather_data"]
        self.runway_data = self.data["runway_data"]

    def test_log_reader(self):
        # Make sure the test workbook is closed.
        log_to_xlsx.log_reader(self.log_file_path,
                               self.name_converter_file_path,
                               self.data_sources_path,
                               self.excel_file_path,
                               self.excel_file_name,
                               self.flight_date,
                               self.flight_number,
                               self.weather_data,
                               self.runway_data)

        def worksheet_tester(sheet, cell, expected_answer):
            """
            A function to check the values of specified cells in a workbook.
            """
            # Opens workbook created from test log file.
            test_workbook = load_workbook(self.base_path +
                                          "/test_xlsx.xlsx", read_only=True)
            # Find the RCIN sheet and a cell (chosen to be mid flight).
            test_worksheet = test_workbook[sheet]
            # Defines the cell in the excel naming convention
            cell = get_column_letter(cell[1] + 1) + str(cell[0] + 1)
            # Checks a cell in the time column (unique).
            test_cell = test_worksheet[cell]
            self.assertEqual(str(test_cell.value), str(expected_answer))

        sheetlist = ["GPS", "RCIN", "BARO", "ARSP", "ATT", "VIBE", "CTUN",
                     "AOA"]
        # These values need to be adjusted if the
        # test_log_to_xlsx.log file gets changed.
        cell = [[699, 13], [3499, 14], [1399, 8], [1449, 9], [3499, 8],
                [3499, 6], [3499, 8], [3499, 2]]
        expected_answer = [156014236, 156314116, 156214156, 161213822,
                           156314116, 156314116, 156314116, 156314116]
        """
        The following loop checks that a value in each .xlsx sheet is correct.
        The chosen values are time values and hence there should only be one
        present per sheet.
        """
        for sheet in range(len(sheetlist)):
            worksheet_tester(sheetlist[sheet], cell[sheet],
                             expected_answer[sheet])

    def test_log_reader_multi(self):
        # Defines relevant variables
        log_file_paths = [self.log_file_path, self.log_file_path]
        name_converter_file_path = self.name_converter_file_path
        data_sources_path = self.data_sources_path
        excel_file_path = self.excel_file_path
        excel_file_names = ["test_xlsx", "test_xlsx_multi"]
        flight_dates = [self.flight_date, "87654321"]
        flight_numbers = [self.flight_number, "6"]
        weather_data_multi = [self.weather_data, {'Action_time_hh:mm': 'DE:FG', 'Pressure_Pa': 'HIJ',
                                                  'Temperature_C': 'KLMN', 'Wind_direction_degrees': 'OP',
                                                  'Wind_speed_mps': 'ABC'}]
        runway_data_multi = [self.runway_data, {'runway_surface': '1234', 'surface_condition': '45',
                                                'take_off_direction': '678'}]

        # Runs the multi-xlsx generator
        log_to_xlsx.log_reader_multi(log_file_paths, name_converter_file_path, data_sources_path,
                                     excel_file_path, excel_file_names, flight_dates, flight_numbers,
                                     weather_data_multi, runway_data_multi)

        # Checks that the second file has been made.
        self.assertTrue(os.path.exists(self.base_path + "test_xlsx_multi.xlsx"))


if __name__ == '__main__':
    unittest.main()
