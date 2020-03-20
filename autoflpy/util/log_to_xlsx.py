
import os
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell



"""
This code converts a .log file (generated using mission planner from a .bin
file) into a .xls document to be used with the automated flight log creator.
@author: Adrian Weishaeupl
aw6g15@soton.ac.uk 2019

Based on work done by Samuel Pearson (sp1g18@soton.ac.uk) (06-08/2019)
"""


def log_reader(log_file_path, name_converter_file_path, data_sources_path,
               excel_file_path, excel_file_name, flight_date, flight_number, weather_data, runway_data):
    """Creates a formatted excel file from a log file. """
    print('Starting log reader')
    print('Creating new work book')
    # Creates a new write only workbook for faster writing
    workbook = Workbook(write_only=True)
    # Opens log file
    print('Reading log file')
    log_opened = open(log_file_path, "r")
    # Reads contents
    log_contents_text = log_opened.read()
    # Closes file
    log_opened.close()
    # Opens file
    name_list_opened = open(name_converter_file_path, "r")
    # Reads contents
    name_list_text = name_list_opened.read()
    # Closes file
    name_list_opened.close()
    # Opens file
    data_sources_opened = open(data_sources_path, "r")
    # Reads contents
    data_sources_text = data_sources_opened.read()
    # Closes file
    data_sources_opened.close()
    # Splits text from data sources into individual lines
    data_sources = data_sources_text.split("\n")[1:]
    # splits name list into lines and ignored the first key line.
    name_list = name_list_text.split("\n")[1:]
    # splits log contents about each new line
    log_contents = log_contents_text.split("\n")

    # Goes through each line
    print('Populating work book')
    for line in log_contents:
        # Splits data into columns
        data = line.split(", ")
        if data[0] == "FMT":
            # Specifies whether data is available
            data_available = False
            # Checks to see if data was recorded for a particular heading.
            for check_line in log_contents:
                # splits data into lines
                check_line_list = check_line.split(", ")
                # Checks through all data to see if there was any data recorded
                # for a particular variable.
                if data[3] == check_line_list[0]:
                    # Specifies that data is available
                    data_available = True
                    # breaks from for loop
                    break
            if data[3] != "FMT" and data[3] != "UNIT" and data[3] != "FMTU" \
                    and data_available is True and data[3] in data_sources:

                # Defines line for titles and resets this for every sheet
                heading_line = []

                # Creates a new worksheet for all of the data.
                worksheet = workbook.create_sheet(title=data[3])

                # Excludes the first time column and puts it at the end.
                data_list_time_end = data[-1].split(",")[1:]
                # Appends time column at the end.
                data_list_time_end.append(data[-1].split(",")[0])
                # Creates the headings and appends the time column to the end
                # to match the format of the previous data sets.
                for heading_name in data_list_time_end:
                    # Code will be here to find the units
                    unit = "unavailable_"
                    # heading name check code will go here.
                    heading = heading_name
                    # Goes through the names in the name_list to check the
                    # units
                    for name_data in name_list:
                        # splits name list
                        name_info = name_data.split(", ")
                        # Checks to see if the information in the list matches
                        # that being from the log file
                        if name_info[0] == data[3] and name_info[1] == \
                                heading_name:
                            # Sets heading name to be same as from name
                            heading = name_info[2]
                            # Checks to see if there were units
                            if unit == "no unit":
                                unit = ""
                            else:
                                # Sets unit to be that from name converter list
                                unit = name_info[3] + "_"
                            break
                    # Creates heading from data
                    heading = heading + "_" + unit + data[3] + "_" + flight_date + "_Flight" + flight_number

                    # Creates the heading line
                    heading_cell = WriteOnlyCell(worksheet, value=heading)
                    heading_line.append(heading_cell)

                # Writes heading line to the worksheet
                worksheet.append(heading_line)

                # Goes through all data searching for a match
                for lines in log_contents:
                    # Splits line data
                    line_data = lines.split(", ")
                    #  Checks to see if data name is the one being searched.
                    if line_data[0] == data[3]:
                        # Resets row to be written
                        row = []
                        # Goes through all data in line, starting from the
                        # column after the time column
                        for recorded_data in line_data[2:]:
                            row.append(recorded_data)
                        # Appends time data for the row
                        row.append(line_data[1])
                        # Writes the row to the worksheet
                        worksheet.append(row)
        else:
            # Ends for loop and so saves code
            break

    # Adds custom weather data to the xlsx document
    worksheet = workbook.create_sheet("WEATHER_DATA")
    weather_keys = list(weather_data.keys())
    weather_values = list(weather_data.values())
    worksheet.append(weather_keys)
    worksheet.append(weather_values)

    # Adds custom runway data to the xlsx document
    worksheet = workbook.create_sheet("RUNWAY_DATA")
    runway_keys = list(runway_data.keys())
    runway_values = list(runway_data.values())
    worksheet.append(runway_keys)
    worksheet.append(runway_values)

    # Saves file
    print('Saving workbook')
    xlsx_file_name_and_path = excel_file_path + os.sep + excel_file_name + ".xlsx"
    workbook.save(filename=xlsx_file_name_and_path)
    print('Log reader finished for {}'.format(str(excel_file_name)))


def log_reader_multi(log_file_paths, name_converter_file_path, data_sources_path,
                     excel_file_path, excel_file_names, flight_dates, flight_numbers, weather_data_multi,
                     runway_data_multi):
    """Runs the log_reader for once per flight log entered in the Input_file.json"""
    # Iterates through the number of flights
    for flight in range(len(flight_numbers)):
        print("Creating workbook for {}".format(str(excel_file_names[flight])))
        log_reader(log_file_paths[flight], name_converter_file_path, data_sources_path, excel_file_path,
                   excel_file_names[flight], flight_dates[flight], flight_numbers[flight], weather_data_multi[flight],
                   runway_data_multi[flight])
